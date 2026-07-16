from fastapi import APIRouter, Depends, HTTPException, status
from pydantic import BaseModel
from sqlalchemy import delete, func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

import io
import re
from datetime import datetime

import openpyxl
from fastapi import UploadFile
from fastapi.responses import StreamingResponse

from app.db.session import get_db
from app.deps import CurrentUser, get_current_user, require_role, scope_class_filter
from app.models.attendance import AttendanceRecord
from app.models.fee_voucher import FeeVoucher
from app.models.extra_charge import ExtraCharge
from app.models.grade import Grade
from app.models.payment_history import PaymentHistory
from app.models.student import Student
from app.schemas.student import StudentCreate, StudentOut, StudentSearchParams, StudentUpdate
from app.schemas.student_import import (
    AnalyzeResponse,
    ExecuteRequest,
    ExecuteResult,
    PreviewRequest,
    PreviewResponse,
)
from app.services import student_import_service as import_service
from app.services.fee_status import aggregate_fee_status, fee_summaries_for_students
from app.services.search import fuzzy_pick, looks_like_name, text_search_condition
from app.schemas.fee_voucher import VoucherOut
from app.schemas.extra_charge import ChargeOut

router = APIRouter(prefix="/api/students", tags=["students"])

CLASS_SEQUENCE = [
    "Playgroup", "Nursery", "Prep",
    "Grade 1", "Grade 2", "Grade 3", "Grade 4", "Grade 5",
    "Grade 6", "Grade 7", "Grade 8", "Grade 9", "Grade 10",
]


class PromoteRequest(BaseModel):
    increments: dict[str, float] = {}


def _generate_registration_no(db: Session) -> str:
    last = db.execute(
        select(Student.registration_no).order_by(Student.student_id.desc()).limit(1)
    ).scalar_one_or_none()
    if last is None:
        return "REG-0001"
    digits = "".join(ch for ch in last if ch.isdigit())
    next_num = (int(digits) + 1) if digits else 1
    return f"REG-{next_num:04d}"


def _fee_status(db: Session, student_id: int) -> str:
    vouchers = db.execute(
        select(FeeVoucher).where(FeeVoucher.student_id == student_id)
    ).scalars().all()
    return aggregate_fee_status(vouchers)


def _norm_phone(p: str | None) -> str:
    """Digits-only form of a phone number, keeping the last 10 (national) digits
    so different formattings of the same number match."""
    digits = re.sub(r"\D", "", p or "")
    return digits[-10:] if len(digits) >= 10 else digits


class SiblingOut(BaseModel):
    student_id: int
    name: str
    class_name: str
    total_pending: float


@router.get("/{student_id}/siblings", response_model=list[SiblingOut])
def get_siblings(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Other students sharing this student's phone number (i.e. siblings), each
    with their total remaining dues (pending fee vouchers + extra charges)."""
    student = db.get(Student, student_id)
    if student is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Student not found")
    core = _norm_phone(student.phone)
    if not core:
        return []
    others = db.execute(
        select(Student).where(Student.student_id != student_id)
    ).scalars().all()
    siblings = [s for s in others if _norm_phone(s.phone) == core]
    if not siblings:
        return []
    summaries = fee_summaries_for_students(db, [s.student_id for s in siblings])
    return [
        SiblingOut(
            student_id=s.student_id,
            name=s.name,
            class_name=s.class_name,
            total_pending=float(summaries[s.student_id][1]),
        )
        for s in siblings
    ]


def _total_pending(db: Session, student_id: int) -> float:
    voucher_pending = db.execute(
        select(
            (FeeVoucher.total_amount - FeeVoucher.paid_amount - FeeVoucher.discount_amount).label("p")
        ).where(FeeVoucher.student_id == student_id)
    ).scalars().all()
    voucher_pending = [max(float(v), 0) for v in voucher_pending]
    charge_pending = db.execute(
        select(ExtraCharge.remaining_amount).where(
            ExtraCharge.student_id == student_id, ExtraCharge.status != "Paid"
        )
    ).scalars().all()
    charge_pending = [float(c) for c in charge_pending]
    return round(sum(voucher_pending) + sum(charge_pending), 2)


def _to_out(db: Session, student: Student, current_user: CurrentUser | None = None) -> StudentOut:
    out = StudentOut.model_validate(student)
    if current_user is not None and current_user.role_name == "Teacher":
        # Teachers get the roster (for attendance) but never fee figures —
        # use /students/pending-fee-names for a names-only pending flag instead.
        out.fee_status = None
        out.total_pending = None
        out.default_fee = None
        return out
    out.fee_status = _fee_status(db, student.student_id)
    out.total_pending = _total_pending(db, student.student_id)
    return out


def _to_out_many(db: Session, students: list[Student], current_user: CurrentUser | None = None) -> list[StudentOut]:
    """List variant of _to_out: two fee queries for the whole result set
    instead of three per student."""
    if current_user is not None and current_user.role_name == "Teacher":
        return [_to_out(db, s, current_user) for s in students]
    summaries = fee_summaries_for_students(db, [s.student_id for s in students])
    outs = []
    for s in students:
        out = StudentOut.model_validate(s)
        out.fee_status, out.total_pending = summaries[s.student_id]
        outs.append(out)
    return outs


@router.get("", response_model=list[StudentOut])
def list_students(
    search: str = "",
    class_filter: str = "",
    status_filter: str = "",
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    # Teachers default to their own assigned class but may request any class
    # (they need the roster to take any class's attendance); Admin/Accountant
    # see whatever class_filter they pass.
    class_filter = scope_class_filter(current_user, class_filter) or ""
    query = select(Student)
    if search:
        query = query.where(
            text_search_condition(search, Student.name, Student.registration_no, Student.cnic)
        )
    if class_filter:
        query = query.where(Student.class_name == class_filter)
    if status_filter:
        query = query.where(Student.status == status_filter)
    query = query.order_by(Student.student_id.desc())
    students = list(db.execute(query).scalars().all())
    # Typo-tolerant supplement: surface close spelling variants of a name that
    # the exact/normalized match above missed (e.g. "abdulhady" → "Abdul Hadi").
    if search and looks_like_name(search):
        students = _augment_with_fuzzy(db, search, students, class_filter, status_filter)
    return _to_out_many(db, students, current_user)


def _augment_with_fuzzy(
    db: Session, search: str, base: list[Student], class_filter: str, status_filter: str,
) -> list[Student]:
    """Append fuzzy name matches (respecting the same class/status filters) to
    the exact matches, keeping exact matches first."""
    candidate_query = select(Student.student_id, Student.name)
    if class_filter:
        candidate_query = candidate_query.where(Student.class_name == class_filter)
    if status_filter:
        candidate_query = candidate_query.where(Student.status == status_filter)
    pairs = db.execute(candidate_query).all()
    extra_ids = fuzzy_pick(search, pairs, exclude_ids={s.student_id for s in base})
    if not extra_ids:
        return base
    extra = db.execute(select(Student).where(Student.student_id.in_(extra_ids))).scalars().all()
    order = {sid: i for i, sid in enumerate(extra_ids)}
    extra.sort(key=lambda s: order.get(s.student_id, 0))
    return base + list(extra)


@router.get("/class-counts")
def get_class_counts(db: Session = Depends(get_db)):
    """Active student count per class, in promotion order — used by the
    Promote Students screen to show how many students sit in each class."""
    counts = dict(
        db.execute(
            select(Student.class_name, func.count())
            .where(Student.status == "Active")
            .group_by(Student.class_name)
        ).all()
    )
    return [{"class_name": c, "count": counts.get(c, 0)} for c in CLASS_SEQUENCE]


# ---------------------------------------------------------------- bulk import
MAX_IMPORT_UPLOAD_BYTES = 20 * 1024 * 1024  # 20MB


@router.post("/import/analyze", response_model=AnalyzeResponse, dependencies=[Depends(require_role("Admin"))])
async def analyze_import_file(file: UploadFile, db: Session = Depends(get_db)):
    """Step 1-3 of the import wizard: parses the uploaded spreadsheet once and
    returns every row as plain column->value dicts plus a suggested mapping.
    All later steps (preview/execute) operate on this same row data sent back
    in the request body — nothing is kept on the server between requests, so
    this fits a stateless/ephemeral-filesystem host with no extra work."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Please upload an .xlsx, .xls, or .csv file.")
    content = await file.read(MAX_IMPORT_UPLOAD_BYTES + 1)
    if len(content) > MAX_IMPORT_UPLOAD_BYTES:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "File is too large (max 20MB).")

    try:
        if file.filename.lower().endswith(".csv"):
            import csv
            text = content.decode("utf-8-sig")
            reader = csv.reader(text.splitlines())
            sheet_rows = list(reader)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb["Students"] if "Students" in wb.sheetnames else wb.active
            sheet_rows = list(ws.iter_rows(values_only=True))
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Could not read this file — make sure it's a valid spreadsheet export.")

    if not sheet_rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "This file has no rows.")

    columns = [str(h).strip() if h is not None else f"Column {i + 1}" for i, h in enumerate(sheet_rows[0])]
    raw_rows = [dict(zip(columns, row)) for row in sheet_rows[1:] if any(v not in (None, "") for v in row)]

    suggested_mapping = import_service.suggest_column_mapping(columns)
    known_classes = db.execute(select(Grade.class_name)).scalars().all()

    class_col = next((c for c, f in suggested_mapping.items() if f == "class_name"), None)
    distinct_class_values = sorted({str(r[class_col]).strip() for r in raw_rows if class_col and r.get(class_col)})
    suggested_class_mapping = import_service.suggest_class_mapping(distinct_class_values, known_classes)

    return AnalyzeResponse(
        columns=columns,
        suggested_mapping=suggested_mapping,
        raw_rows=raw_rows,
        total_rows=len(raw_rows),
        distinct_class_values=distinct_class_values,
        suggested_class_mapping=suggested_class_mapping,
        known_classes=known_classes,
    )


@router.post("/import/preview", response_model=PreviewResponse, dependencies=[Depends(require_role("Admin"))])
def preview_import(payload: PreviewRequest, db: Session = Depends(get_db)):
    rows = import_service.build_preview_rows(payload.raw_rows, payload.mapping, payload.class_value_mapping)
    valid = sum(1 for r in rows if r.status == "valid")
    invalid = sum(1 for r in rows if r.status == "invalid")
    duplicate = sum(1 for r in rows if r.status == "duplicate")
    missing = sum(1 for r in rows if r.missing_fields)
    return PreviewResponse(
        total_rows=len(rows), valid_rows=valid, invalid_rows=invalid,
        duplicate_rows=duplicate, missing_fields_rows=missing, rows=rows,
    )


@router.post("/import/execute", response_model=ExecuteResult, dependencies=[Depends(require_role("Admin"))])
def execute_import(payload: ExecuteRequest, db: Session = Depends(get_db)):
    if payload.import_mode == "delete_all" and not payload.confirm_delete_all:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Set confirm_delete_all=true to proceed with this mode.")
    return import_service.execute_import(
        db, payload.raw_rows, payload.mapping, payload.class_value_mapping,
        payload.import_mode, payload.only_valid_rows,
    )


# ---------------------------------------------------------------- bulk export
_EXPORT_COLUMNS = [
    ("registration_no", "Registration Number"), ("name", "Student Name"), ("father_name", "Father's Name"),
    ("class_name", "Class"), ("phone", "Parent Phone"), ("cnic", "CNIC"), ("default_fee", "Monthly Fee"),
    ("dob", "Date of Birth"), ("admission_date", "Admission Date"), ("b_form", "B-Form Number"),
    ("address", "Address"), ("status", "Status"),
]


@router.get("/export", dependencies=[Depends(require_role("Admin", "Accountant"))])
def export_students_bulk(
    format: str = "xlsx",
    scope: str = "all",
    student_ids: str = "",
    search: str = "",
    class_filter: str = "",
    status_filter: str = "",
    db: Session = Depends(get_db),
):
    """Exports students in the same column layout the import wizard expects,
    so a round-trip export->edit->import works with zero remapping. `scope`
    picks which students: all / selected (uses student_ids) / filtered (uses
    class_filter/status_filter) / search (uses search)."""
    query = select(Student)
    if scope == "selected":
        ids = [int(x) for x in student_ids.split(",") if x.strip().isdigit()]
        if not ids:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "No students selected.")
        query = query.where(Student.student_id.in_(ids))
    elif scope == "filtered":
        if class_filter:
            query = query.where(Student.class_name == class_filter)
        if status_filter:
            query = query.where(Student.status == status_filter)
    elif scope == "search":
        if search:
            query = query.where(
                text_search_condition(search, Student.name, Student.registration_no, Student.cnic)
            )
    query = query.order_by(Student.student_id)
    students = db.execute(query).scalars().all()

    rows = [[getattr(s, key) for key, _ in _EXPORT_COLUMNS] for s in students]
    headers = [label for _, label in _EXPORT_COLUMNS]
    filename_base = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if format == "csv":
        import csv
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        writer.writerows(rows)
        out = io.BytesIO(buffer.getvalue().encode("utf-8-sig"))
        return StreamingResponse(
            out, media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename_base}.csv"},
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(headers)
    for row in rows:
        ws.append([str(v) if v is not None else "" for v in row])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename_base}.xlsx"},
    )


@router.post("/promote-all", dependencies=[Depends(require_role("Admin"))])
def promote_all_students(payload: PromoteRequest, db: Session = Depends(get_db)):
    """Promotes every active student to the next class in CLASS_SEQUENCE,
    adding that class's fee increment to their default_fee. Students
    currently in the last class (Grade 10) are marked Inactive instead of
    promoted further.

    Since promotion marks the end of an academic year, it also prunes
    history that's no longer needed for every active student (promoted or
    graduating alike): fully-paid fee vouchers/extra charges are deleted
    (pending ones are kept so nothing owed is lost), and all attendance rows
    are deleted outright. PaymentHistory is untouched — it's a financial
    audit trail with no FK to vouchers/charges, so it stays intact even
    after the voucher/charge it paid for is gone."""
    grades = {g.class_name: g for g in db.execute(select(Grade)).scalars().all()}
    grade_fees = {name: float(g.fee_amount) for name, g in grades.items()}
    students = db.execute(select(Student).where(Student.status == "Active")).scalars().all()
    student_ids = [s.student_id for s in students]

    by_class: dict[str, list[Student]] = {}
    for s in students:
        by_class.setdefault(s.class_name, []).append(s)

    for class_name in CLASS_SEQUENCE:
        if class_name not in grades:
            db.add(Grade(class_name=class_name, fee_amount=0))
            grades[class_name] = None
    db.flush()

    promoted = 0
    deactivated = 0
    for i, class_name in enumerate(CLASS_SEQUENCE):
        bucket = by_class.get(class_name, [])
        if not bucket:
            continue
        increment = float(payload.increments.get(class_name, 0) or 0)
        for s in bucket:
            base_fee = float(s.default_fee) if s.default_fee is not None else grade_fees.get(class_name, 0)
            new_fee = base_fee + increment
            if i == len(CLASS_SEQUENCE) - 1:
                s.status = "Inactive"
                deactivated += 1
            else:
                s.class_name = CLASS_SEQUENCE[i + 1]
                s.default_fee = new_fee
                promoted += 1

    if student_ids:
        db.execute(delete(FeeVoucher).where(FeeVoucher.student_id.in_(student_ids), FeeVoucher.status == "Paid"))
        db.execute(delete(ExtraCharge).where(ExtraCharge.student_id.in_(student_ids), ExtraCharge.status == "Paid"))
        db.execute(delete(AttendanceRecord).where(AttendanceRecord.student_id.in_(student_ids)))

    db.commit()
    return {"promoted": promoted, "deactivated": deactivated}


@router.get("/pending-fee-names", dependencies=[Depends(require_role("Admin", "Accountant", "Teacher"))])
def get_pending_fee_names(
    class_name: str = "",
    db: Session = Depends(get_db),
):
    """Teacher-safe view: just the names of active students in a class who
    have any pending fee voucher or extra charge — no amounts, no other
    financial detail. Any role permitted here may query any class."""
    if not class_name:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "class_name is required")

    students = db.execute(
        select(Student).where(Student.class_name == class_name, Student.status == "Active")
    ).scalars().all()
    student_ids = [s.student_id for s in students]

    with_pending_vouchers = set(
        db.execute(
            select(FeeVoucher.student_id)
            .where(FeeVoucher.student_id.in_(student_ids), FeeVoucher.status != "Paid")
            .distinct()
        ).scalars().all()
    ) if student_ids else set()
    with_pending_charges = set(
        db.execute(
            select(ExtraCharge.student_id)
            .where(ExtraCharge.student_id.in_(student_ids), ExtraCharge.status != "Paid")
            .distinct()
        ).scalars().all()
    ) if student_ids else set()

    return [
        {"student_id": s.student_id, "name": s.name, "registration_no": s.registration_no}
        for s in students
        if s.student_id in with_pending_vouchers or s.student_id in with_pending_charges
    ]


@router.get("/{student_id}", response_model=StudentOut)
def get_student(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    student = db.get(Student, student_id)
    if student is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Student not found")
    scope_class_filter(current_user, student.class_name)
    return _to_out(db, student, current_user)


@router.post("", response_model=StudentOut, dependencies=[Depends(require_role("Admin", "Accountant"))])
def add_student(payload: StudentCreate, db: Session = Depends(get_db)):
    # Registration numbers come from a read-increment-write, so two admissions
    # submitted at the same moment can collide on the unique constraint —
    # retry with a fresh number instead of surfacing a 500.
    for attempt in range(3):
        student = Student(registration_no=_generate_registration_no(db), **payload.model_dump())
        db.add(student)
        try:
            db.commit()
            return _to_out(db, student)
        except IntegrityError as exc:
            db.rollback()
            if "registration_no" not in str(exc.orig) or attempt == 2:
                raise
    raise HTTPException(status.HTTP_409_CONFLICT, "Could not allocate a registration number — please try again.")


@router.put("/{student_id}", response_model=StudentOut, dependencies=[Depends(require_role("Admin", "Accountant"))])
def update_student(student_id: int, payload: StudentUpdate, db: Session = Depends(get_db)):
    student = db.get(Student, student_id)
    if student is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Student not found")
    for field, value in payload.model_dump().items():
        setattr(student, field, value)
    db.commit()
    return _to_out(db, student)


@router.patch("/{student_id}/status", dependencies=[Depends(require_role("Admin", "Accountant"))])
def set_status(student_id: int, new_status: str, db: Session = Depends(get_db)):
    student = db.get(Student, student_id)
    if student is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Student not found")
    student.status = new_status
    db.commit()
    return _to_out(db, student)


@router.delete("/{student_id}", dependencies=[Depends(require_role("Admin"))])
def delete_student(student_id: int, db: Session = Depends(get_db)):
    """Permanently deletes a student and every record tied to them: fee
    vouchers (and their QR codes), extra charges, attendance history, and
    contacts all cascade at the database level (ondelete=CASCADE on
    student_id/voucher_id). PaymentHistory has no FK — it's keyed by
    target_type/target_id — so it's purged explicitly here for this
    student's vouchers and charges before the student row is deleted."""
    student = db.get(Student, student_id)
    if student is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Student not found")

    voucher_ids = db.execute(
        select(FeeVoucher.voucher_id).where(FeeVoucher.student_id == student_id)
    ).scalars().all()
    if voucher_ids:
        db.execute(
            delete(PaymentHistory).where(
                PaymentHistory.target_type == "fee_voucher",
                PaymentHistory.target_id.in_(voucher_ids),
            )
        )

    charge_ids = db.execute(
        select(ExtraCharge.charge_id).where(ExtraCharge.student_id == student_id)
    ).scalars().all()
    if charge_ids:
        db.execute(
            delete(PaymentHistory).where(
                PaymentHistory.target_type == "extra_charge",
                PaymentHistory.target_id.in_(charge_ids),
            )
        )

    db.delete(student)
    db.commit()
    return {"detail": "Deleted"}


@router.post("/search-advanced", response_model=list[StudentOut], dependencies=[Depends(require_role("Admin", "Accountant"))])
def search_advanced(
    params: StudentSearchParams,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    class_filter = scope_class_filter(current_user, params.class_filter)
    query = select(Student)
    if params.name:
        query = query.where(text_search_condition(params.name, Student.name))
    if params.registration_no:
        query = query.where(text_search_condition(params.registration_no, Student.registration_no))
    if params.cnic:
        query = query.where(text_search_condition(params.cnic, Student.cnic))
    if params.phone:
        query = query.where(text_search_condition(params.phone, Student.phone))
    if params.father_name:
        query = query.where(text_search_condition(params.father_name, Student.father_name))
    if params.address:
        query = query.where(text_search_condition(params.address, Student.address))
    if class_filter:
        query = query.where(Student.class_name == class_filter)
    if params.status_filter:
        query = query.where(Student.status == params.status_filter)
    if params.admission_year:
        query = query.where(Student.admission_date.cast(str).like(f"{params.admission_year}%"))
    if params.dob_from:
        query = query.where(Student.dob >= params.dob_from)
    if params.dob_to:
        query = query.where(Student.dob <= params.dob_to)
    query = query.order_by(Student.student_id.desc())
    students = db.execute(query).scalars().all()
    results = _to_out_many(db, students)

    if params.fee_status_filter:
        results = [s for s in results if s.fee_status == params.fee_status_filter]
    return results


@router.get("/{student_id}/vouchers", response_model=list[VoucherOut], dependencies=[Depends(require_role("Admin", "Accountant"))])
def get_student_vouchers(student_id: int, db: Session = Depends(get_db)):
    return db.execute(
        select(FeeVoucher).where(FeeVoucher.student_id == student_id).order_by(FeeVoucher.fee_month_sort)
    ).scalars().all()


@router.get("/{student_id}/charges", response_model=list[ChargeOut], dependencies=[Depends(require_role("Admin", "Accountant"))])
def get_student_charges(student_id: int, db: Session = Depends(get_db)):
    return db.execute(
        select(ExtraCharge).where(ExtraCharge.student_id == student_id).order_by(ExtraCharge.created_at.desc())
    ).scalars().all()
