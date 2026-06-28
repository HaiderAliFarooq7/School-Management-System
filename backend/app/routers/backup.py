import asyncio
import io
import os
import tempfile
from datetime import datetime

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from starlette.background import BackgroundTask
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.db.session import engine, get_db
from app.deps import require_role
from app.logging_config import logger
from app.models.extra_charge import ExtraCharge
from app.models.fee_voucher import FeeVoucher
from app.models.notification_log import NotificationLog
from app.models.notification_queue import NotificationQueue
from app.models.payment_history import PaymentHistory
from app.models.qr_code import QRCode
from app.models.student import Student
from app.models.student_contact import StudentContact
from app.schemas.backup import ResetRequest
from app.services.backup_service import backup_database, backup_filename, restore_database

router = APIRouter(
    prefix="/api/backup", tags=["backup"], dependencies=[Depends(require_role("Admin"))]
)

STUDENT_COLUMNS = [
    "student_id", "registration_no", "name", "father_name", "class_name", "dob",
    "admission_date", "b_form", "cnic", "phone", "address", "status", "default_fee",
]

MAX_DUMP_UPLOAD_BYTES = 200 * 1024 * 1024  # 200MB — generous for a school DB dump
MAX_EXCEL_UPLOAD_BYTES = 20 * 1024 * 1024  # 20MB — generous for a student-list spreadsheet


async def _save_upload_capped(file: UploadFile, dest_path: str, max_bytes: int) -> None:
    """Streams an upload to disk in chunks instead of loading it into memory
    in one go, aborting once `max_bytes` is exceeded — protects against a
    huge upload exhausting memory on a small hosting instance."""
    written = 0
    with open(dest_path, "wb") as f:
        while chunk := await file.read(1024 * 1024):
            written += len(chunk)
            if written > max_bytes:
                raise HTTPException(
                    status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                    f"File is too large (max {max_bytes // (1024 * 1024)}MB).",
                )
            f.write(chunk)


@router.get("/database")
def run_backup():
    """Streams a fresh pg_dump straight to the admin's browser, then deletes
    the temp file from the server — no backup is ever stored on the host."""
    try:
        path = backup_database()
    except Exception as exc:
        logger.exception("Database backup failed")
        raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, "Backup failed. Check server logs for details.")
    return FileResponse(
        path,
        media_type="application/octet-stream",
        filename=backup_filename(),
        background=BackgroundTask(lambda: os.path.exists(path) and os.remove(path)),
    )


@router.post("/restore")
async def restore_backup(file: UploadFile, db: Session = Depends(get_db)):
    """Restores the database from an uploaded .dump file. The upload is
    written to a temp file, restored with pg_restore, and deleted immediately
    afterwards — nothing is retained on the server.

    Before running pg_restore we close this request's own DB session and
    dispose the whole connection pool. Without this, the auth check that ran
    for this very request (require_role) leaves a session "idle in
    transaction" holding a read lock on user_account — and pg_restore's
    `--clean` needs an exclusive lock on that same table to drop its
    constraints, so it hangs forever waiting on a lock held by the request
    that's waiting on it. Running the actual restore in a thread keeps the
    (necessarily blocking) subprocess call from freezing the whole server's
    single event loop for every other request while it runs."""
    if not file.filename or not file.filename.lower().endswith(".dump"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Please upload a .dump backup file.")
    fd, tmp_path = tempfile.mkstemp(suffix=".dump", prefix="sms_restore_")
    os.close(fd)
    try:
        await _save_upload_capped(file, tmp_path, MAX_DUMP_UPLOAD_BYTES)
        db.close()
        engine.dispose()
        await asyncio.to_thread(restore_database, tmp_path)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Database restore failed")
        raise HTTPException(
            status.HTTP_500_INTERNAL_SERVER_ERROR,
            "Restore failed — make sure this is a valid backup file. Check server logs for details.",
        )
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    return {"detail": "Database restored successfully."}


@router.get("/export-students.xlsx")
def export_students(db: Session = Depends(get_db)):
    students = db.execute(select(Student).order_by(Student.student_id)).scalars().all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(STUDENT_COLUMNS)
    for s in students:
        ws.append([getattr(s, c) for c in STUDENT_COLUMNS])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/import-students")
async def import_students(file: UploadFile, db: Session = Depends(get_db)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Please upload an .xlsx file.")
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", prefix="sms_import_")
    os.close(fd)
    try:
        await _save_upload_capped(file, tmp_path, MAX_EXCEL_UPLOAD_BYTES)
        wb = openpyxl.load_workbook(tmp_path)
    except HTTPException:
        raise
    except Exception:
        logger.exception("Student import file could not be read")
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Could not read this file — make sure it's a valid .xlsx export.")
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    ws = wb["Students"] if "Students" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0}

    headers = [str(h).strip() for h in rows[0]]
    count = 0
    for raw in rows[1:]:
        record = dict(zip(headers, raw))
        reg_no = record.get("registration_no")
        if not reg_no or not record.get("name"):
            continue
        existing = db.execute(
            select(Student).where(Student.registration_no == reg_no)
        ).scalar_one_or_none()
        fields = {
            "name": record.get("name") or "",
            "father_name": record.get("father_name") or "",
            "class_name": record.get("class_name") or "",
            "dob": record.get("dob") or None,
            "admission_date": record.get("admission_date") or None,
            "b_form": record.get("b_form") or None,
            "cnic": record.get("cnic") or None,
            "phone": record.get("phone") or None,
            "address": record.get("address") or None,
            "status": record.get("status") or "Active",
            "default_fee": record.get("default_fee") or None,
        }
        if existing:
            for k, v in fields.items():
                setattr(existing, k, v)
        else:
            db.add(Student(registration_no=reg_no, **fields))
        count += 1
    db.commit()
    return {"imported": count}


@router.post("/reset")
def reset_database(payload: ResetRequest, db: Session = Depends(get_db)):
    """Clears all operational data (students, vouchers, charges, history,
    contacts, notifications, QR codes) but leaves school settings, roles,
    and user accounts intact. Requires an explicit confirm flag."""
    if not payload.confirm:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Set confirm=true to proceed with reset.")

    for model in (QRCode, NotificationQueue, NotificationLog, StudentContact, PaymentHistory, ExtraCharge, FeeVoucher, Student):
        db.execute(delete(model))
    db.commit()
    return {"detail": "All operational data has been reset."}
